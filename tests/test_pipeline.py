import pandas as pd
import pytest
from pipeline import compute_macro_result
import io
from pipeline import parse_dpt
from pipeline import zone_for_filename


def test_compute_macro_result_exact_match_points():
    """When the spectrum contains exact wavenumber matches, returns the formula's value."""
    df = pd.DataFrame({
        0: [1000.0, 1500.0, 1590.0, 1800.0, 2000.0, 2242.0, 2500.0],
        1: [0.10,   0.20,   0.50,   0.40,   0.30,   0.80,   0.05],
    })
    result = compute_macro_result(df, wn_a=1590, wn_b=2242)
    # z1 = 0.5 (row at 1590), z2 = 0.8 (row at 2242)
    expected = (0.29 * 0.5) / ((0.29 * 0.5) + 0.8) * 100
    assert result == pytest.approx(expected, abs=1e-9)


def test_compute_macro_result_picks_closest_when_no_exact_match():
    """When wavenumbers don't appear exactly, picks the row with smallest abs difference."""
    df = pd.DataFrame({
        0: [1585.0, 1592.0, 2240.0, 2245.0],  # no exact 1590 or 2242
        1: [0.10,   0.50,   0.80,   0.20],
    })
    # Closest to 1590 is 1592 (|diff|=2), closest to 2242 is 2240 (|diff|=2)
    result = compute_macro_result(df, wn_a=1590, wn_b=2242)
    expected = (0.29 * 0.50) / ((0.29 * 0.50) + 0.80) * 100
    assert result == pytest.approx(expected, abs=1e-9)


def test_compute_macro_result_default_wavenumbers():
    """Defaults are wn_a=1590, wn_b=2242."""
    df = pd.DataFrame({
        0: [1590.0, 2242.0],
        1: [1.0,    1.0],
    })
    result = compute_macro_result(df)
    expected = (0.29 * 1.0) / ((0.29 * 1.0) + 1.0) * 100
    assert result == pytest.approx(expected, abs=1e-9)


def test_parse_dpt_comma_separated_no_header():
    """Comma-separated numeric data with no header row."""
    content = b"1000.5,0.123\n1001.0,0.124\n1002.0,0.125\n"
    df = parse_dpt(content)
    assert df.shape == (3, 2)
    assert df.iloc[0, 0] == pytest.approx(1000.5)
    assert df.iloc[0, 1] == pytest.approx(0.123)


def test_parse_dpt_tab_separated_no_header():
    """Tab-separated numeric data with no header row."""
    content = b"1000.5\t0.123\n1001.0\t0.124\n"
    df = parse_dpt(content)
    assert df.shape == (2, 2)
    assert df.iloc[1, 0] == pytest.approx(1001.0)


def test_parse_dpt_whitespace_separated():
    """Whitespace-separated numeric data."""
    content = b"1000.5 0.123\n1001.0 0.124\n"
    df = parse_dpt(content)
    assert df.shape == (2, 2)


def test_parse_dpt_with_text_header():
    """First row is text → treat as header and skip."""
    content = b"wavenumber,intensity\n1000.5,0.123\n1001.0,0.124\n"
    df = parse_dpt(content)
    assert df.shape == (2, 2)
    assert df.iloc[0, 0] == pytest.approx(1000.5)


def test_parse_dpt_utf16_bom():
    """UTF-16 with BOM is decoded correctly."""
    text = "1000.5,0.123\n1001.0,0.124\n"
    content = text.encode("utf-16")  # adds BOM
    df = parse_dpt(content)
    assert df.shape == (2, 2)
    assert df.iloc[0, 0] == pytest.approx(1000.5)


def test_parse_dpt_too_small_returns_none():
    """Files under 10 bytes return None (skip signal)."""
    assert parse_dpt(b"1,2") is None


def test_parse_dpt_real_fixture():
    """Parses an actual .dpt fixture file end-to-end."""
    from pathlib import Path
    fixture_dir = Path(__file__).parent / "fixtures" / "sample_dpt"
    dpt_files = sorted(fixture_dir.glob("*.dpt"))
    assert len(dpt_files) >= 1, "Fixture .dpt files missing — re-run Task 2"
    content = dpt_files[0].read_bytes()
    df = parse_dpt(content)
    assert df is not None
    assert df.shape[1] == 2
    assert df.shape[0] > 100  # FTIR spectra are typically thousands of points


def test_zone_for_filename_standard_forms():
    """Recognizes 'zone N', 'zoneN', 'zN', etc. case-insensitively."""
    assert zone_for_filename("Sample Zone 1.dpt") == 1
    assert zone_for_filename("sample_zone2_a.dpt") == 2
    assert zone_for_filename("data Zone3.dpt") == 3
    assert zone_for_filename("DATA ZONE 4.dpt") == 4
    assert zone_for_filename("file z5 thing.dpt") == 5
    assert zone_for_filename("Z6_sample.dpt") == 6


def test_zone_for_filename_decimal_variants():
    """Recognizes 'zone N.' forms (e.g., 'Zone 4.0')."""
    assert zone_for_filename("Sample Zone 4.0.dpt") == 4


def test_zone_for_filename_no_zone_returns_none():
    """Returns None when no zone pattern matches."""
    assert zone_for_filename("random_file.dpt") is None
    assert zone_for_filename("EXTRACT_sample.dpt") is None


def test_zone_for_filename_skips_merged_files():
    """Files with 'merged' in the name are skipped (return None)."""
    assert zone_for_filename("ZONE_1_merged.xlsx") is None
    assert zone_for_filename("zone_2_MERGED.dpt") is None


from pipeline import merge_zone
from openpyxl import load_workbook


def test_merge_zone_two_files_column_layout():
    """Per-file: 8 columns appended. Plus trailing mean column at the end."""
    df_a = pd.DataFrame({
        0: [1590.0, 2242.0, 1595.0, 2243.0],
        1: [0.5,    0.8,    0.5,    0.8],
    })
    df_b = pd.DataFrame({
        0: [1590.0, 2242.0, 1595.0, 2243.0],
        1: [1.0,    2.0,    1.0,    2.0],
    })
    files = [("file_a.dpt", df_a), ("file_b.dpt", df_b)]
    merged = merge_zone(files)
    # 2 files * 8 columns each + 1 trailing mean = 17 columns
    assert merged.shape[1] == 17

    # For each file, result_l gets 2 values appended (macro + python).
    # With this rigged data, macro_result (at 1590/2242) == python_result (at 1595/2243)
    # because all four wavenumbers map to distinct rows but the formula
    # uses the exact-match intensities.
    # df_a: macro uses rows 0 (0.5) and 1 (0.8); python uses rows 2 (0.5) and 3 (0.8) — same formula result.
    # df_b: macro uses rows 0 (1.0) and 1 (2.0); python uses rows 2 (1.0) and 3 (2.0) — same.
    r_a = (0.29 * 0.5) / ((0.29 * 0.5) + 0.8) * 100
    r_b = (0.29 * 1.0) / ((0.29 * 1.0) + 2.0) * 100
    expected_mean = (r_a + r_a + r_b + r_b) / 4
    assert float(merged.iloc[0, -1]) == pytest.approx(expected_mean, abs=1e-9)


def test_merge_zone_empty_returns_empty():
    """Zero files → empty DataFrame."""
    merged = merge_zone([])
    assert merged.empty


from pipeline import process_dpt_files


def test_process_dpt_files_with_4_unzoned_files_treats_as_zone_1(tmp_path):
    """If exactly 4 non-merged files have no zone name, all become Zone 1."""
    content = b"1590,0.5\n2242,0.8\n1595,0.5\n2243,0.8\n"
    files = [(f"sample_{i}.dpt", content) for i in range(4)]
    filename, xlsx_bytes = process_dpt_files(files)
    assert filename.endswith("_FINAL_OUTPUT.xlsx")
    assert xlsx_bytes is not None and len(xlsx_bytes) > 0

    out_path = tmp_path / "out.xlsx"
    out_path.write_bytes(xlsx_bytes)
    wb = load_workbook(out_path)
    assert "ZONE 1" in wb.sheetnames


def test_process_dpt_files_groups_by_zone_name():
    """Files named with zone numbers are grouped accordingly."""
    content = b"1590,0.5\n2242,0.8\n1595,0.5\n2243,0.8\n"
    files = [
        ("sample_zone1_a.dpt", content),
        ("sample_zone1_b.dpt", content),
        ("sample_zone2_a.dpt", content),
        ("sample_zone2_b.dpt", content),
    ]
    filename, xlsx_bytes = process_dpt_files(files)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert "ZONE 1" in wb.sheetnames
    assert "ZONE 2" in wb.sheetnames


def test_process_dpt_files_progress_callback_called():
    """Progress callback receives status messages during processing."""
    content = b"1590,0.5\n2242,0.8\n"
    files = [(f"sample_{i}.dpt", content) for i in range(4)]
    messages: list[str] = []
    process_dpt_files(files, progress_callback=messages.append)
    assert len(messages) > 0
    assert any("zone" in m.lower() or "parse" in m.lower() or "merge" in m.lower() for m in messages)


def test_process_dpt_files_empty_input_raises():
    """No valid files → raises ValueError."""
    with pytest.raises(ValueError):
        process_dpt_files([])


def test_process_dpt_files_all_files_corrupt_raises():
    """All files under 10 bytes → raises ValueError."""
    files = [(f"sample_{i}.dpt", b"x") for i in range(3)]
    with pytest.raises(ValueError):
        process_dpt_files(files)


def test_process_dpt_files_smoke_on_real_fixtures():
    """Run the pipeline on the real .dpt fixtures and verify it produces a valid xlsx."""
    from pathlib import Path
    import io as _io2

    fixture_dir = Path(__file__).parent / "fixtures" / "sample_dpt"
    dpt_paths = sorted(fixture_dir.glob("*.dpt"))
    assert len(dpt_paths) >= 1, "Fixture .dpt files missing"

    files = [(p.name, p.read_bytes()) for p in dpt_paths]
    filename, xlsx_bytes = process_dpt_files(files)

    wb = load_workbook(_io2.BytesIO(xlsx_bytes))
    assert len(wb.sheetnames) >= 1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.max_row > 0
        assert ws.max_column > 0


def _numeric_values(ws):
    """Yield all numeric cell values from a worksheet."""
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, (int, float)):
                yield float(v)


def test_pipeline_matches_golden_output():
    """Numerically compare new-pipeline output to a saved golden xlsx from the original."""
    import io as _io3
    import math
    from pathlib import Path

    fixture_dir = Path(__file__).parent / "fixtures"
    golden_path = fixture_dir / "golden_output.xlsx"
    if not golden_path.exists():
        pytest.skip("Golden output not present — generate by running the original pipeline (see plan Task 9)")

    dpt_paths = sorted((fixture_dir / "sample_dpt").glob("*.dpt"))
    files = [(p.name, p.read_bytes()) for p in dpt_paths]
    _, xlsx_bytes = process_dpt_files(files)

    new_wb = load_workbook(_io3.BytesIO(xlsx_bytes))
    golden_wb = load_workbook(golden_path)

    assert set(new_wb.sheetnames) == set(golden_wb.sheetnames), \
        f"Sheet names differ: new={new_wb.sheetnames} golden={golden_wb.sheetnames}"

    for sheet_name in new_wb.sheetnames:
        new_vals = list(_numeric_values(new_wb[sheet_name]))
        golden_vals = list(_numeric_values(golden_wb[sheet_name]))
        assert len(new_vals) == len(golden_vals), \
            f"Sheet {sheet_name}: cell-count mismatch ({len(new_vals)} vs {len(golden_vals)})"
        for i, (n, g) in enumerate(zip(new_vals, golden_vals)):
            assert math.isclose(n, g, rel_tol=1e-6, abs_tol=1e-6), \
                f"Sheet {sheet_name}, value index {i}: new={n} golden={g}"
