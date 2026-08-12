import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from dat_pipeline import (
    convert_dat_files,
    dat_to_xlsx_bytes,
    detect_column_headers,
    parse_dat,
)

FIXTURE_DIR = Path(__file__).parent / "fixtures" / "sample_dat"


# --- detect_column_headers -------------------------------------------------


def test_detect_named_psi_header():
    """The psi(°) / Intensity / Sigma_I layout is recognized by name."""
    lines = [
        "# EDF_DataBlockID  0.Image.Psd\n",
        "    psi(°)   Intensity(a.u.)   Sigma_I(a.u.)\n",
        "0.0   2160224.96   195977.927\n",
    ]
    headers, start = detect_column_headers(lines)
    assert headers == ["psi(°)", "Intensity(a.u.)", "Sigma_I(a.u.)"]
    assert start == 2  # data begins on the line after the header


def test_detect_named_q_header():
    """The q(A-1) / I(q) / Sig(q) layout is recognized by name."""
    lines = [
        "# DataType  FloatValue\n",
        "q(A-1)   I(q)   Sig(q)\n",
        "0.0035   23.46   2.42\n",
    ]
    headers, start = detect_column_headers(lines)
    assert headers == ["q(A-1)", "I(q)", "Sig(q)"]
    assert start == 2


def test_detect_falls_back_to_three_column_layout():
    """No named header, 3 numeric columns → the q layout."""
    lines = [
        "################\n",
        "# Dim_1   1038\n",
        "1.0 2.0 3.0\n",
    ]
    headers, start = detect_column_headers(lines)
    assert headers == ["q(A-1)", "I(q)", "Sig(q)"]
    assert start == 2  # the numeric line itself is data


def test_detect_falls_back_to_six_column_layout():
    """No named header, 6 numeric columns → psi + q layouts concatenated."""
    lines = ["1.0 2.0 3.0 4.0 5.0 6.0\n"]
    headers, start = detect_column_headers(lines)
    assert headers == [
        "psi(°)",
        "Intensity(a.u.)",
        "Sigma_I(a.u.)",
        "q(A-1)",
        "I(q)",
        "Sig(q)",
    ]
    assert start == 0


def test_detect_falls_back_to_generic_column_names():
    """An unrecognized column count gets generic Column_N names."""
    lines = ["1.0 2.0 3.0 4.0\n"]
    headers, start = detect_column_headers(lines)
    assert headers == ["Column_1", "Column_2", "Column_3", "Column_4"]
    assert start == 0


def test_detect_ignores_comment_lines_when_scanning_for_data():
    """Lines starting with # are not mistaken for numeric data."""
    lines = ["# 1.0 2.0 3.0\n", "5.0 6.0 7.0\n"]
    headers, start = detect_column_headers(lines)
    assert start == 1


# --- parse_dat -------------------------------------------------------------


def test_parse_dat_extracts_metadata_key_value_pairs():
    """'# key value' header lines become metadata entries."""
    content = b"# DataType   FloatValue\n# Dim_1   1038\nq(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n"
    parsed = parse_dat(content)
    assert parsed["metadata"]["DataType"] == "FloatValue"
    assert parsed["metadata"]["Dim_1"] == "1038"


def test_parse_dat_skips_the_banner_line():
    """The ####### banner has no '# ' prefix and yields no metadata entry."""
    content = b"#####################\n# Dim_1   1038\n1.0 2.0 3.0\n"
    parsed = parse_dat(content)
    assert list(parsed["metadata"]) == ["Dim_1"]


def test_parse_dat_collects_numeric_rows():
    """Numeric lines after the header become float rows."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n4.0 5.0 6.0\n"
    parsed = parse_dat(content)
    assert parsed["data"] == [[1.0, 2.0, 3.0], [4.0, 5.0, 6.0]]


def test_parse_dat_ignores_blank_and_nonnumeric_rows():
    """Blank lines, comments, and garbage inside the data block are dropped."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n\n# note\nnot numbers\n4.0 5.0 6.0\n"
    parsed = parse_dat(content)
    assert parsed["data"] == [[1.0, 2.0, 3.0], [4.0, 5.0, 6.0]]


def test_parse_dat_requires_at_least_two_columns():
    """Single-value rows are not data (matches the original's len >= 2 rule)."""
    content = b"q(A-1) I(q) Sig(q)\n1.0\n2.0 3.0\n"
    parsed = parse_dat(content)
    assert parsed["data"] == [[2.0, 3.0]]


def test_parse_dat_empty_content_returns_none():
    """Nothing to parse → None, the skip signal."""
    assert parse_dat(b"") is None


def test_parse_dat_real_psi_fixture():
    """A real psi-format .dat parses end to end."""
    parsed = parse_dat((FIXTURE_DIR / "sample_psi.dat").read_bytes())
    assert parsed["headers"] == ["psi(°)", "Intensity(a.u.)", "Sigma_I(a.u.)"]
    assert len(parsed["data"]) == 20
    assert parsed["data"][0] == [0.0, 2160224.96, 195977.927]
    assert parsed["metadata"]["DataType"] == "FloatValue"


def test_parse_dat_real_q_fixture():
    """A real q-format .dat parses end to end."""
    parsed = parse_dat((FIXTURE_DIR / "sample_q.dat").read_bytes())
    assert parsed["headers"] == ["q(A-1)", "I(q)", "Sig(q)"]
    assert len(parsed["data"]) == 20
    assert parsed["metadata"]["EDF_DataBlockID"] == "0.Image.Psd"
    assert parsed["metadata"]["Dim_1"] == "1038"


# --- dat_to_xlsx_bytes -----------------------------------------------------


def _sheets(xlsx_bytes):
    return load_workbook(io.BytesIO(xlsx_bytes))


def test_xlsx_has_metadata_and_data_sheets():
    """Both sheets are written, with the detected column names as the data header row."""
    parsed = parse_dat((FIXTURE_DIR / "sample_q.dat").read_bytes())
    wb = _sheets(dat_to_xlsx_bytes(parsed))
    assert wb.sheetnames == ["Metadata", "Data"]

    meta = wb["Metadata"]
    assert [c.value for c in meta[1]] == ["Parameter", "Value"]

    data = wb["Data"]
    assert [c.value for c in data[1]] == ["q(A-1)", "I(q)", "Sig(q)"]
    assert data.max_row == 21  # header + 20 data rows


def test_xlsx_pads_ragged_rows():
    """Rows shorter than the widest row are padded with blanks, not dropped."""
    parsed = {
        "metadata": {"k": "v"},
        "data": [[1.0, 2.0, 3.0], [4.0, 5.0]],
        "headers": ["q(A-1)", "I(q)", "Sig(q)"],
    }
    data = _sheets(dat_to_xlsx_bytes(parsed))["Data"]
    assert data.max_column == 3
    assert data.cell(row=3, column=3).value is None


def test_xlsx_extends_headers_when_data_is_wider():
    """More columns than headers → generic names fill the gap."""
    parsed = {
        "metadata": {},
        "data": [[1.0, 2.0, 3.0, 4.0, 5.0]],
        "headers": ["q(A-1)", "I(q)", "Sig(q)"],
    }
    data = _sheets(dat_to_xlsx_bytes(parsed))["Data"]
    assert [c.value for c in data[1]] == [
        "q(A-1)",
        "I(q)",
        "Sig(q)",
        "Column_4",
        "Column_5",
    ]


def test_xlsx_truncates_headers_when_data_is_narrower():
    """Fewer columns than headers → surplus header names are dropped."""
    parsed = {"metadata": {}, "data": [[1.0, 2.0]], "headers": ["a", "b", "c"]}
    data = _sheets(dat_to_xlsx_bytes(parsed))["Data"]
    assert [c.value for c in data[1]] == ["a", "b"]


def test_xlsx_metadata_only_file_still_produces_a_workbook():
    """Header but no numeric data → a Metadata sheet alone is valid output."""
    parsed = {"metadata": {"Dim_1": "1038"}, "data": [], "headers": ["a"]}
    wb = _sheets(dat_to_xlsx_bytes(parsed))
    assert wb.sheetnames == ["Metadata"]


def test_xlsx_returns_none_when_there_is_nothing_to_write():
    """No metadata and no data would close a zero-sheet workbook, which openpyxl
    rejects with 'At least one sheet must be visible'. Return None instead."""
    parsed = {"metadata": {}, "data": [], "headers": []}
    assert dat_to_xlsx_bytes(parsed) is None


# --- convert_dat_files -----------------------------------------------------


def _zip_names(zip_bytes):
    return zipfile.ZipFile(io.BytesIO(zip_bytes)).namelist()


def test_convert_produces_one_xlsx_per_dat():
    """Each input file gets its own workbook inside the zip, keeping its stem."""
    files = [
        (p.name, p.read_bytes()) for p in sorted(FIXTURE_DIR.glob("*.dat"))
    ]
    zip_name, zip_bytes = convert_dat_files(files)
    assert sorted(_zip_names(zip_bytes)) == ["sample_psi.xlsx", "sample_q.xlsx"]


def test_convert_zip_name_derives_from_first_file():
    """The download is named after the first uploaded file's stem."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n"
    zip_name, _ = convert_dat_files([("run_001.dat", content)])
    assert zip_name == "run_001_CONVERTED.zip"


def test_convert_suffixes_duplicate_filenames():
    """Same-named files from different folders must not collide inside the zip."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n"
    files = [("scan.dat", content), ("scan.dat", content), ("scan.dat", content)]
    _, zip_bytes = convert_dat_files(files)
    assert sorted(_zip_names(zip_bytes)) == [
        "scan.xlsx",
        "scan_2.xlsx",
        "scan_3.xlsx",
    ]


def test_convert_skips_unusable_files_but_keeps_the_batch():
    """One empty file is skipped; the valid ones still convert."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n"
    files = [("good.dat", content), ("empty.dat", b""), ("also_good.dat", content)]
    _, zip_bytes = convert_dat_files(files)
    assert sorted(_zip_names(zip_bytes)) == ["also_good.xlsx", "good.xlsx"]


def test_convert_reports_progress_and_a_final_summary():
    """The callback receives per-file lines plus an N converted / M skipped tail."""
    content = b"q(A-1) I(q) Sig(q)\n1.0 2.0 3.0\n"
    messages = []
    convert_dat_files(
        [("good.dat", content), ("empty.dat", b"")],
        progress_callback=messages.append,
    )
    assert any("good.dat" in m for m in messages)
    assert any("empty.dat" in m and "kip" in m for m in messages)
    assert any("1 converted" in m and "1 skipped" in m for m in messages)


def test_convert_no_files_raises():
    """An empty upload list is a caller error."""
    with pytest.raises(ValueError):
        convert_dat_files([])


def test_convert_all_files_unusable_raises():
    """Nothing converted → ValueError, so the UI can show an error not a zip."""
    with pytest.raises(ValueError):
        convert_dat_files([("a.dat", b""), ("b.dat", b"")])


def test_convert_output_workbooks_open_cleanly():
    """Every workbook in the zip is readable and non-empty."""
    files = [(p.name, p.read_bytes()) for p in sorted(FIXTURE_DIR.glob("*.dat"))]
    _, zip_bytes = convert_dat_files(files)
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    for name in zf.namelist():
        wb = load_workbook(io.BytesIO(zf.read(name)))
        assert "Data" in wb.sheetnames
        assert wb["Data"].max_row > 1
